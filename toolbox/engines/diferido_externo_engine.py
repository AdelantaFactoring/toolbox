"""
⚙️ Diferido Externo Engine V2
LÓGICA COPIADA EXACTA DE DiferidoExternoCalcular V1 
"""

import pandas as pd
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import re
from io import BytesIO

try:
    from ..core.base_engine import BaseEngine
except ImportError:
    raise ImportError(
        "DiferidoExternoEngine V2 requiere BaseEngine de imports relativos"
    )

# CONSTANTES COPIADAS EXACTAS DE V1
MONTH_MAP = {
    "ENE": "enero",
    "FEB": "febrero",
    "MAR": "marzo",
    "ABR": "abril",
    "MAY": "mayo",
    "JUN": "junio",
    "JUL": "julio",
    "AGO": "agosto",
    "SET": "septiembre",
    "OCT": "octubre",
    "NOV": "noviembre",
    "DIC": "diciembre",
}


def convert_date_col(col: str) -> str:
    """FUNCIÓN COPIADA EXACTA DE V1"""
    m = re.match(r"^(\d{4})([A-Za-z]+)$", col)
    if m:
        year = m.group(1)
        month_abbr = m.group(2).upper()
        month_full = MONTH_MAP.get(month_abbr, month_abbr.lower())
        return f"{month_full}-{year}"
    return col


def reorder_date_columns(date_cols: list[str]) -> list[str]:
    """FUNCIÓN COPIADA EXACTA DE V1"""

    def sort_key(col: str):
        try:
            # separa por guión: ej: "enero-2023"
            month_str, year_str = col.split("-")
            year = int(year_str)
            months_order = [
                "enero",
                "febrero",
                "marzo",
                "abril",
                "mayo",
                "junio",
                "julio",
                "agosto",
                "septiembre",
                "octubre",
                "noviembre",
                "diciembre",
            ]
            month_idx = (
                months_order.index(month_str) if month_str in months_order else 99
            )
            return (year, month_idx)
        except Exception:
            return (9999, 99)

    return sorted(date_cols, key=sort_key)


class DiferidoExternoEngine(BaseEngine):
    """Motor que contiene TODA la lógica de DiferidoExternoCalcular V1"""

    def __init__(self):
        super().__init__()

    def read_excel_file(
        self, file_path: str | BytesIO, sheet_name: str, usecols: str, date_col: str
    ):
        """COPIADO EXACTO DE V1"""
        # Si file_path es un objeto file-like (tiene método 'read')
        if hasattr(file_path, "read"):
            file_path.seek(0)
            df = pd.read_excel(
                file_path, sheet_name=sheet_name, skiprows=1, usecols=usecols
            )
        else:
            # Asumimos que es una ruta de archivo
            df = pd.read_excel(
                file_path, sheet_name=sheet_name, skiprows=1, usecols=usecols
            )
        # Eliminar columnas sin nombre (generadas como 'Unnamed')
        df = df.loc[:, ~df.columns.str.contains("^Unnamed")]
        df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        return df

    def last_day_of_month(self, date: datetime) -> datetime:
        """COPIADO EXACTO DE V1"""
        if date.month == 12:
            next_month = date.replace(year=date.year + 1, month=1, day=1)
        else:
            next_month = date.replace(month=date.month + 1, day=1)
        return next_month - timedelta(days=1)

    def auto_get_usecols(
        self,
        file_path: str | BytesIO,
        sheet_name: str,
        fixed_range: tuple[str, str],
        stop_marker: str,
    ) -> tuple[str, list[str]]:
        """COPIADO EXACTO DE V1"""
        # Si el archivo está cerrado, reabrirlo usando su nombre (si existe)
        if hasattr(file_path, "closed") and file_path.closed:
            if hasattr(file_path, "name"):
                file_path = open(file_path.name, "rb")
            else:
                raise ValueError("No se puede reabrir el archivo: no tiene 'name'.")
        else:
            if hasattr(file_path, "seek"):
                file_path.seek(0)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        # Verifica que la hoja exista
        if sheet_name not in wb.sheetnames:
            raise KeyError(
                f"Worksheet '{sheet_name}' does not exist. Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        fixed_start = column_index_from_string(fixed_range[0])
        fixed_end = column_index_from_string(fixed_range[1])

        # Definir dónde inicia el grupo de fechas según la hoja:
        if sheet_name.upper() == "US":
            date_start = column_index_from_string("S")
        elif sheet_name.upper() == "MN":
            date_start = column_index_from_string("N")
        else:
            date_start = fixed_end + 1

        end_month = None
        raw_date_cols = []
        # Iterar desde date_start hasta el final o hasta encontrar el stop_marker en la fila 2.
        for col_idx in range(date_start, ws.max_column + 1):
            cell_value = ws.cell(row=2, column=col_idx).value
            # Si se encuentra el stop_marker, se finaliza el grupo de fechas
            if cell_value is not None and cell_value == stop_marker:
                end_month = col_idx - 1
                break
            cell1 = ws.cell(row=1, column=col_idx).value
            combined = ""
            if cell1:
                combined += str(cell1).strip()
            if cell_value:
                combined += str(cell_value).strip()
            raw_date_cols.append(combined)
        if end_month is None:
            end_month = ws.max_column

        # Obtener las letras correspondientes al rango de fechas
        date_cols_letters = [
            get_column_letter(i) for i in range(date_start, end_month + 1)
        ]
        # Obtener las letras para el rango fijo
        fixed_cols = [get_column_letter(i) for i in range(fixed_start, fixed_end + 1)]
        usecols_str = ",".join(fixed_cols + date_cols_letters)

        # Procesar raw_date_cols para obtener los nombres de fechas
        # Filtrar solo aquellos que cumplan el patrón: 4 dígitos seguidos de al menos 3 letras
        raw_filtered = [
            col
            for col in raw_date_cols
            if re.match(r"^\d{4}[A-Za-z]{3,}$", col, re.IGNORECASE)
        ]
        # Convertir al formato deseado, ej.: "2021AGO" -> "agosto-2021"
        converted = [convert_date_col(col) for col in raw_filtered]
        # Ordenar cronológicamente
        converted = reorder_date_columns(converted)

        return usecols_str, converted

    def process_excel_files(
        self,
        file_path: str | BytesIO,
        pen_end_date: str = "2025-07-31",
        usd_end_date: str = "2025-12-31",
        read_type: str | None = None,
    ):
        """COPIADO EXACTO DE V1"""
        # PARA PEN:
        # Se obtiene el usecols y las columnas de fecha a partir del rango fijo ("B", "H")
        pen_usecols, pen_date_cols = self.auto_get_usecols(
            file_path=file_path,
            sheet_name="MN",
            fixed_range=("B", "H"),
            stop_marker="VAL",
        )
        print(f"Usecols PEN: {pen_usecols}")
        # print(f"Usecols PEN: {pen_usecols}")
        df_pen = self.read_excel_file(
            file_path=file_path,
            sheet_name="MN",
            usecols=pen_usecols,
            date_col="Fecha de Op",
        )
        print(
            f"dataframe df_pen: {df_pen.loc[df_pen["LIQUIDACIÓN"] == "LIQ2502000076", :]}"
        )
        print(f"Columnas PEN: {pen_date_cols}")
        # Calcular el número de columnas fijas en PEN (de "B" a "H")
        fixed_pen_count = (
            column_index_from_string("H") - column_index_from_string("B") + 1
        )
        print(f"Columnas fijas PEN: {fixed_pen_count}")
        fixed_pen = list(df_pen.columns[:fixed_pen_count])
        print(f"Columnas fijas PEN: {fixed_pen}")
        expected_pen_cols = fixed_pen + pen_date_cols
        df_pen = df_pen.iloc[:, : len(expected_pen_cols)]
        df_pen.columns = expected_pen_cols
        df_pen["Moneda"] = "PEN"
        print(
            f"dataframe df_pen final: {df_pen.loc[df_pen["LIQUIDACIÓN"] == "LIQ2502000076", :]}"
        )

        # PARA USD:
        # Se obtiene el usecols y las columnas de fecha a partir del rango fijo ("B", "G")
        usd_usecols, usd_date_cols = self.auto_get_usecols(
            file_path=file_path,
            sheet_name="US",
            fixed_range=("B", "G"),
            stop_marker="VAL",
        )

        df_usd = self.read_excel_file(
            file_path=file_path,
            sheet_name="US",
            usecols=usd_usecols,
            date_col="Fecha de Op",
        )

        # Calcular el número de columnas fijas en USD (de "B" a "G")
        fixed_usd_count = (
            column_index_from_string("G") - column_index_from_string("B") + 1
        )
        fixed_usd = list(df_usd.columns[:fixed_usd_count])
        print(f"Columnas fijas USD: {fixed_usd}")
        expected_usd_cols = fixed_usd + usd_date_cols
        df_usd = df_usd.iloc[:, : len(expected_usd_cols)]
        df_usd.columns = expected_usd_cols
        df_usd["Moneda"] = "USD"

        df = pd.concat([df_pen, df_usd], axis=0).rename(
            columns={
                "LIQUIDACIÓN": "CodigoLiquidacion",
                "Fecha de Op": "FechaOperacion",
                "Días Efect": "DiasEfectivo",
                "Intereses sin IGV": "Interes",
                "F.Pago Confirmada / Real": "FechaConfirmado",
            }
        )
        # print(f"dataframe df: {df.loc[df["CodigoLiquidacion"] == "LIQ2502000076", :]}")
        df["CodigoLiquidacion"] = df["CodigoLiquidacion"].str.strip()
        df["N° de Factura referencial"] = df["N° de Factura referencial"].str.strip()
        df.rename(columns={"N° de Factura referencial": "NroDocumento"}, inplace=True)
        df = df.drop(columns=["TIPO DE OPERACIÓN"], errors="ignore")
        return df

    def replace_columns_with_dates(self, df: pd.DataFrame):
        """COPIADO EXACTO DE V1"""
        # Columnas fijas deseadas en el orden correcto
        columns_to_keep = [
            "CodigoLiquidacion",
            "NroDocumento",
            "FechaOperacion",
            "FechaConfirmado",
            "Moneda",
            "Interes",
            "DiasEfectivo",
        ]
        # Excluir las columnas indeseadas
        unwanted = ["Día de fecha de Op.", "Día de fecha de Pago"]
        # Tomar el resto de columnas (asumidas de fecha) sin las que no deseamos
        date_columns = [
            col
            for col in df.columns
            if col not in columns_to_keep and col not in unwanted
        ]
        date_columns = reorder_date_columns(date_columns)
        new_columns = columns_to_keep + date_columns
        return df[new_columns]

    def calcular_diferido_externo(
        self, file_path: str | BytesIO, hasta: str
    ) -> pd.DataFrame:
        """LÓGICA PRINCIPAL COPIADA EXACTA DE V1"""
        # Validamos el formato
        if not hasta or not re.match(r"^\d{4}-\d{2}$", hasta):
            raise ValueError(
                "El parámetro 'hasta' es obligatorio y debe tener el formato 'YYYY-MM'."
            )

        # Procesar el archivo Excel y obtener el DataFrame con los datos mensuales
        df = self.process_excel_files(file_path)

        # Convertir 'hasta' a fecha y obtener su último día
        hasta_date = datetime.strptime(hasta, "%Y-%m")
        hasta_date = self.last_day_of_month(hasta_date)

        # Filtrar según FechaOperacion
        df = df[df["FechaOperacion"] <= hasta_date]

        # Reorganizar columnas y limpiar DataFrame
        df = self.replace_columns_with_dates(df)
        df = df.dropna(subset=["CodigoLiquidacion"])
        # df = df.dropna(subset=["NroDocumento"])
        df = df.fillna(0)
        return df
